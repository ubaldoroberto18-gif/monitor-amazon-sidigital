# -*- coding: utf-8 -*-
# ============================================================
#  MONITOR AMAZON MX - SI DIGITAL v8 GitHub Actions
#  + Nombres reales de competidores
#  + Elegibilidad Buy Box por ASIN
# ============================================================

import os, time, smtplib, urllib.request, re
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
from datetime import datetime
from openpyxl import load_workbook
from sp_api.api import Products, CatalogItems
from sp_api.base import Marketplaces

REFRESH_TOKEN     = os.environ["REFRESH_TOKEN"]
LWA_APP_ID        = os.environ["LWA_APP_ID"]
LWA_CLIENT_SECRET = os.environ["LWA_CLIENT_SECRET"]
MI_SELLER_ID      = os.environ["SELLER_ID"]
GMAIL_USER        = os.environ["GMAIL_USER"]
GMAIL_PASSWORD    = "".join(os.environ["GMAIL_PASSWORD"].split())
CORREO_DESTINO    = "roberto.flores@sidigital.com.mx"
ARCHIVO_EXCEL     = "Analisis_Amazon_SIDigital.xlsx"
ESPERA            = 5

credentials_sp = dict(
    refresh_token=REFRESH_TOKEN,
    lwa_app_id=LWA_APP_ID,
    lwa_client_secret=LWA_CLIENT_SECRET,
)

# Cache de nombres para no repetir consultas
NOMBRES_CACHE = {"": "Sin dato", MI_SELLER_ID: "SI DIGITAL (tú)"}

def num(x):
    try: return float(x)
    except: return None

def find_header_row(ws, busca="ASIN", max_scan=12):
    for r in range(1, max_scan + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if v and str(v).strip() == busca:
                return r
    return 3

def col_map(ws, header_row):
    h = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if v:
            h[str(v).strip()] = c
    return h

def safe_set(ws, row, col, value):
    try:
        ws.cell(row=row, column=col).value = value
    except AttributeError:
        pass

def nombre_vendedor(seller_id):
    """Intenta obtener el nombre comercial del vendedor desde Amazon MX."""
    if not seller_id:
        return "Sin dato"
    if seller_id in NOMBRES_CACHE:
        return NOMBRES_CACHE[seller_id]
    try:
        url = f"https://www.amazon.com.mx/sp?seller={seller_id}"
        req = urllib.request.Request(url, headers={
            "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36",
            "Accept-Language": "es-MX,es;q=0.9",
        })
        html = urllib.request.urlopen(req, timeout=8).read().decode("utf-8", "ignore")
        m = (re.search(r'"sellerName"\s*:\s*"([^"]+)"', html) or
             re.search(r'<h1[^>]*class="[^"]*a-size-large[^"]*"[^>]*>([^<]+)<', html) or
             re.search(r'Vendido\s+por\s*<[^>]+>([^<]+)<', html))
        nombre = m.group(1).strip() if m else seller_id
        NOMBRES_CACHE[seller_id] = nombre
        time.sleep(1)
        return nombre
    except Exception:
        NOMBRES_CACHE[seller_id] = seller_id
        return seller_id

def get_offers(asin):
    for intento in range(4):
        try:
            client = Products(marketplace=Marketplaces.MX, credentials=credentials_sp)
            r = client.get_item_offers(asin, "New")
            p = r.payload or {}
            summary = p.get("Summary", {})
            offers  = p.get("Offers", [])

            # Número de vendedores
            num_off = None
            for c in summary.get("NumberOfOffers", []):
                if c.get("condition") == "new":
                    num_off = c.get("OfferCount")
            if num_off is None:
                num_off = len(offers)

            # Elegibilidad Buy Box
            bb_eligible_count = 0
            for e in summary.get("BuyBoxEligibleOffers", []):
                if e.get("condition") == "new":
                    bb_eligible_count = e.get("OfferCount", 0)
            yo_eligible = "Sin dato"

            # Precio Buy Box
            bb_price = None
            for bb in summary.get("BuyBoxPrices", []):
                bb_price = num(bb.get("ListingPrice", {}).get("Amount"))
                break

            bb_seller_id = "Sin dato"
            yo_bb = "No"
            precios = []
            vendedores = []

            for o in offers:
                precio_o  = num(o.get("ListingPrice", {}).get("Amount"))
                envio_o   = num(o.get("Shipping", {}).get("Amount")) or 0
                total_o   = (precio_o or 0) + envio_o
                logistica = "FBA" if o.get("IsFulfilledByAmazon") else "FBM"
                calif     = o.get("SellerFeedbackRating", {}).get("SellerPositiveFeedbackRating", "—")
                es_bb     = o.get("IsBuyBoxWinner", False)
                sid       = o.get("SellerId", "")
                bb_elig_o = "Sí" if o.get("IsBuyBoxWinner") or o.get("IsFeaturedMerchant") else "No"

                if precio_o:
                    precios.append(total_o)
                if es_bb:
                    bb_seller_id = sid or "Otro vendedor"
                    if sid == MI_SELLER_ID:
                        yo_bb = "Si"
                        yo_eligible = "Sí"
                if sid == MI_SELLER_ID:
                    yo_eligible = bb_elig_o

                vendedores.append({
                    "precio": precio_o, "envio": envio_o, "total": total_o,
                    "logistica": logistica, "calif": calif,
                    "es_bb": "Sí" if es_bb else "No",
                    "seller_id": sid,
                    "bb_eligible": bb_elig_o,
                })

            p_min = min(precios) if precios else None
            p_max = max(precios) if precios else None
            if bb_price is None and precios:
                bb_price = min(precios)

            return ("OK", {
                "bb_price": bb_price, "min": p_min, "max": p_max,
                "num": num_off, "bb_seller_id": bb_seller_id,
                "yo_bb": yo_bb, "yo_eligible": yo_eligible,
                "vendedores": vendedores,
            })
        except Exception as e:
            msg = str(e)
            if "QuotaExceeded" in msg:
                espera = 15 * (intento + 1)
                print(f"(cuota, espero {espera}s)")
                time.sleep(espera)
                continue
            if "InvalidInput" in msg or "invalid ASIN" in msg:
                return ("INVALIDO", None)
            return ("ERROR", str(e)[:80])
    return ("ERROR", "QuotaExceeded tras 4 intentos")

def get_catalog(asin):
    try:
        client = CatalogItems(marketplace=Marketplaces.MX, credentials=credentials_sp)
        r = client.get_catalog_item(asin, marketplaceIds=["A1AM78C64UM0Y8"],
                                    includedData=["salesRanks", "summaries"])
        p = r.payload or {}
        sales = p.get("salesRanks", [{}])
        bsr = subcat = "—"
        if sales:
            for s in sales:
                ranks = s.get("ranks", [])
                if ranks:
                    bsr = ranks[0].get("rank", "—")
                    subcat = ranks[0].get("title", "—")
                    break
        return ("OK", {"bsr": bsr, "subcat": subcat})
    except Exception as e:
        return ("ERROR", str(e)[:60])

def enviar_correo(archivo, ok, sindato, noencontrado, ahora, resumen_competidores):
    msg = MIMEMultipart()
    msg["From"]    = GMAIL_USER
    msg["To"]      = CORREO_DESTINO
    msg["Subject"] = f"Monitor Amazon MX — SI Digital [{ahora}]"

    cuerpo = f"""Hola Roberto,

Aquí está el reporte de monitoreo de precios Amazon MX.

RESUMEN:
- Productos consultados: {ok + sindato + noencontrado}
- Con precio OK: {ok}
- Sin dato / revisar: {sindato}
- ASIN inválidos en MX: {noencontrado}

COMPETIDORES DETECTADOS (quienes tienen la Buy Box):
{resumen_competidores}

El archivo Excel completo va adjunto con el desglose por producto.

— Sistema automático SI Digital
"""
    msg.attach(MIMEText(cuerpo, "plain"))

    with open(archivo, "rb") as f:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header("Content-Disposition",
                        'attachment; filename="Analisis_Amazon_SIDigital.xlsx"')
        msg.attach(part)

    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
        server.login(GMAIL_USER, GMAIL_PASSWORD)
        server.sendmail(GMAIL_USER, CORREO_DESTINO, msg.as_string())
    print(f"Correo enviado a {CORREO_DESTINO}")

def main():
    ahora = datetime.now().strftime("%Y-%m-%d %H:%M")
    print(f"\nMonitoreo Amazon MX — SI Digital v8  [{ahora}]")
    print("="*55)

    wb = load_workbook(ARCHIVO_EXCEL)
    salida = ARCHIVO_EXCEL.replace(".xlsx", "_actualizado.xlsx")

    ws1 = wb["1. Precios y Buy Box"]
    ws2 = wb["2. Competencia"]
    ws3 = wb["3. Elegibilidad BB y FBA"]
    ws4 = wb["4. Catálogo del Producto"]
    ws5 = wb["5. Mis Precios vs Mercado"]
    ws0 = wb["0. Dashboard"]

    HR1 = find_header_row(ws1); H1 = col_map(ws1, HR1)
    HR2 = find_header_row(ws2); H2 = col_map(ws2, HR2)
    HR3 = find_header_row(ws3); H3 = col_map(ws3, HR3)
    HR4 = find_header_row(ws4); H4 = col_map(ws4, HR4)
    HR5 = find_header_row(ws5); H5 = col_map(ws5, HR5)

    col_asin = H1.get("ASIN", 2)
    productos = []
    for row in range(HR1 + 1, ws1.max_row + 1):
        asin = ws1.cell(row=row, column=col_asin).value
        if asin and not str(asin).startswith("#"):
            nombre     = ws1.cell(row=row, column=H1.get("Nombre corto", 3)).value
            precio_pub = ws1.cell(row=row, column=H1.get("Mi precio pub ($)", 11)).value
            costo      = ws1.cell(row=row, column=H1.get("Mi costo ($)", 12)).value
            productos.append((row, str(asin).strip(), nombre, precio_pub, costo))

    total = len(productos)
    ok_count = sindato = noencontrado = 0
    fila_comp = HR2 + 1
    competidores_bb = {}  # nombre -> cuántos productos les ganamos

    for r in range(HR2 + 1, ws2.max_row + 1):
        for c in range(1, 13):
            safe_set(ws2, r, c, None)

    print(f"Productos a consultar: {total}\n")

    for idx, (row, asin, nombre, precio_pub, costo) in enumerate(productos):
        print(f"  [{idx+1}/{total}] {asin} — {str(nombre)[:25]}")
        status, data = get_offers(asin)
        time.sleep(ESPERA)
        safe_set(ws1, row, H1["Fecha revisión"], ahora)

        if status == "INVALIDO":
            safe_set(ws1, row, H1["Estado"],  "Cerrado")
            safe_set(ws1, row, H1["Estatus"], "No encontrado")
            safe_set(ws1, row, H1["Alerta"],  "Listing no encontrado")
            safe_set(ws1, row, H1["Notas"],   "ASIN no válido en Amazon MX")
            noencontrado += 1
            print("       → ASIN inválido en MX")

        elif status == "ERROR":
            safe_set(ws1, row, H1["Estatus"], "Revisar manual")
            safe_set(ws1, row, H1["Alerta"],  "Revisar manual")
            safe_set(ws1, row, H1["Notas"],   f"Error: {data}")
            sindato += 1
            print(f"       → error: {data}")

        else:
            d = data
            bb = d["bb_price"]; num_v = d["num"]; yo_bb = d["yo_bb"]
            yo_elig = d["yo_eligible"]

            # Obtener nombre del vendedor con Buy Box
            bb_nombre = nombre_vendedor(d["bb_seller_id"])
            if bb_nombre != "SI DIGITAL (tú)" and bb_nombre != "Sin dato":
                competidores_bb[bb_nombre] = competidores_bb.get(bb_nombre, 0) + 1

            if bb:
                ant = num(ws1.cell(row=row, column=H1.get("Buy Box ($)", 5)).value)
                sug = round(bb * 0.97, 2)
                cambio = ("Primera revisión" if ant is None
                          else f"Bajó ${ant-bb:,.2f}" if bb < ant
                          else f"Subió ${bb-ant:,.2f}" if bb > ant else "Sin cambio")
                alerta = ("✅ Tienes Buy Box" if yo_bb == "Si"
                          else "🔴 Sin Buy Box" if num_v else "⚪ Sin competencia")

                safe_set(ws1, row, H1["Estado"],           "Activo")
                safe_set(ws1, row, H1["Buy Box ($)"],      bb)
                safe_set(ws1, row, H1["Precio mín ($)"],   d["min"])
                safe_set(ws1, row, H1["Precio máx ($)"],   d["max"])
                safe_set(ws1, row, H1["# Vendedores"],     num_v)
                safe_set(ws1, row, H1["Yo tengo BB"],      yo_bb)
                safe_set(ws1, row, H1["Quién tiene BB"],   bb_nombre)
                safe_set(ws1, row, H1["Precio sug −3% ($)"], sug)
                safe_set(ws1, row, H1["Cambio vs ant."],   cambio)
                safe_set(ws1, row, H1["Estatus"],          "OK")
                safe_set(ws1, row, H1["Alerta"],           alerta)
                safe_set(ws1, row, H1["Notas"],            f"{num_v} vendedor(es)")

                if costo and num(costo):
                    safe_set(ws1, row, H1.get("Margen sug (%)", 14),
                             round((sug - num(costo)) / sug, 4))

                if precio_pub and bb:
                    dif_pct = round((num(precio_pub) - bb) / bb, 4)
                    comp = "Sí" if abs(dif_pct) <= 0.05 else "No"
                    accion = ("⬇️ Bajar — estás por arriba del mercado" if dif_pct > 0.10
                              else "⬆️ Subir — tienes margen disponible" if dif_pct < -0.10
                              else "✅ Precio competitivo — mantener" if comp == "Sí"
                              else "🔍 Ajuste menor recomendado")
                    safe_set(ws5, row, H5.get("Buy Box ($)", 7), bb)
                    safe_set(ws5, row, H5.get("Precio sug −3% ($)", 8), sug)
                    safe_set(ws5, row, H5.get("Dif. mi precio vs BB ($)", 9), round(num(precio_pub)-bb, 0))
                    safe_set(ws5, row, H5.get("Dif. mi precio vs BB (%)", 10), dif_pct)
                    safe_set(ws5, row, H5.get("¿Soy competitivo?", 11), comp)
                    safe_set(ws5, row, H5.get("Precio mín mercado ($)", 12), d["min"])
                    safe_set(ws5, row, H5.get("Precio máx mercado ($)", 13), d["max"])
                    safe_set(ws5, row, H5.get("# Vendedores", 14), num_v)
                    safe_set(ws5, row, H5.get("Acción recomendada", 15), accion)

                # Competencia con nombres
                for v in d["vendedores"]:
                    vend_nombre = nombre_vendedor(v["seller_id"])
                    safe_set(ws2, fila_comp, 1, asin)
                    safe_set(ws2, fila_comp, 2, str(nombre)[:30])
                    safe_set(ws2, fila_comp, 3, v["precio"])
                    safe_set(ws2, fila_comp, 4, v["envio"])
                    safe_set(ws2, fila_comp, 5, v["total"])
                    safe_set(ws2, fila_comp, 6, v["logistica"])
                    safe_set(ws2, fila_comp, 7, v["calif"])
                    safe_set(ws2, fila_comp, 8, v["es_bb"])
                    safe_set(ws2, fila_comp, 9, "Nuevo")
                    safe_set(ws2, fila_comp, 10, ahora)
                    safe_set(ws2, fila_comp, 11, vend_nombre)
                    safe_set(ws2, fila_comp, 12, v["bb_eligible"])
                    fila_comp += 1

                # Elegibilidad
                safe_set(ws3, row, H3.get("Elegible BB", 6), yo_elig)
                safe_set(ws3, row, H3.get("Fecha revisión", 10), ahora)

                ok_count += 1
                print(f"       → OK  BuyBox ${bb:,.2f}  ({num_v} vend.)  BB: {yo_bb}  Elig: {yo_elig}  Gana: {bb_nombre}")
            else:
                safe_set(ws1, row, H1["Estatus"], "Sin dato")
                safe_set(ws1, row, H1["Alerta"],  "⚪ Sin oferta")
                safe_set(ws1, row, H1["Notas"],   "Sin oferta nueva activa")
                sindato += 1
                print("       → sin dato")

        if status == "OK":
            cs, cd = get_catalog(asin)
            time.sleep(ESPERA)
            if cs == "OK":
                safe_set(ws4, row, H4.get("BSR (ranking categoría)", 6), cd["bsr"])
                safe_set(ws4, row, H4.get("Subcategoría", 5), cd["subcat"])
                safe_set(ws4, row, H4.get("Fecha captura", 13), ahora)

    try:
        ws0.cell(row=5, column=1).value = total
        ws0.cell(row=5, column=3).value = ok_count
        ws0.cell(row=2, column=1).value = f"Última actualización: {ahora}"
    except Exception:
        pass

    # Resumen de competidores para el correo
    if competidores_bb:
        resumen = "\n".join([f"  • {nombre}: gana Buy Box en {n} producto(s)"
                             for nombre, n in sorted(competidores_bb.items(),
                             key=lambda x: x[1], reverse=True)])
    else:
        resumen = "  No se detectaron competidores con Buy Box"

    wb.save(salida)
    print(f"\nArchivo guardado: {salida}")
    print("Enviando correo...")
    enviar_correo(salida, ok_count, sindato, noencontrado, ahora, resumen)

    print("\n" + "="*55)
    print(f"Con precio (OK):      {ok_count}")
    print(f"Sin dato / revisar:   {sindato}")
    print(f"No encontrado (MX):   {noencontrado}")
    print(f"Competidores BB detectados: {len(competidores_bb)}")
    print("✅ Proceso completado")

if __name__ == "__main__":
    main()
