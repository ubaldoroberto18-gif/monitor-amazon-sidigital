#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
clasificador_correos.py
========================
Proyecto 3 — SI Digital: Clasificador Automático de Correos.

VERSIÓN 2 — Lee desde Gmail (vía reenviadores), no desde el servidor
propio de sidigital.com.mx, porque ese servidor bloquea las conexiones
que vienen de servidores en la nube como GitHub Actions.

Cómo funciona el puente:
  roberto.flores@sidigital.com.mx  --(reenviador de cPanel)--> analisissidigital.2026@gmail.com
  mercado-libre@sidigital.com.mx   --(reenviador de cPanel)--> analisissidigital.2026@gmail.com

Como los dos llegan al mismo Gmail, el script detecta de cuál cuenta
venía cada correo revisando los encabezados que el reenvío conserva
(Delivered-To / X-Original-To / To).

Qué hace, en orden:
  1. Se conecta por IMAP a UNA sola bandeja: analisissidigital.2026@gmail.com.
  2. Por cada correo, detecta si originalmente era de la cuenta
     principal o de Mercado Libre (mirando los encabezados).
  3. Los de la cuenta principal se clasifican en: Licitaciones,
     Clientes/Proveedores, Otros.
  4. Para los de Licitaciones (alertas de licitary.mx / Compras MX):
     - Parsea la tabla que ya viene en el correo (gratis, sin IA).
     - Deduce el Estado a partir del nombre de la unidad compradora.
     - Aplica un filtro barato (palabras clave + estado) para decidir
       si vale la pena un análisis a fondo.
     - Si sí y hay un PDF directo, lo descarga, extrae el texto y le
       pide a Gemini (gemini-3.1-flash-lite) que responda las 8
       preguntas de análisis, en JSON.
     - Agrega o actualiza filas en el Excel de licitaciones.
  5. Los correos de Mercado Libre: conteo simple por tipo de
     notificación (pregunta, venta, reclamo, mensaje). ESTO ES
     TEMPORAL: en el Paso 4 se reemplaza por la API oficial de
     Mercado Libre, que da datos mucho más completos.
  6. Arma un solo correo "digest" con todo lo anterior y lo envía por
     Gmail (SMTP) a roberto.flores@sidigital.com.mx.

Todo se configura con variables de entorno (vienen de GitHub Secrets
cuando corre en GitHub Actions). Nada de contraseñas ni keys en este
archivo.
"""

import os
import re
import ssl
import json
import imaplib
import smtplib
import email
from email.header import decode_header
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from datetime import datetime, timedelta

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from pypdf import PdfReader
from google import genai

# ===========================================================
# CONFIGURACIÓN (todo viene de GitHub Secrets / variables de entorno)
# ===========================================================

# Gmail es el puente: aquí llegan las copias reenviadas de las 2
# cuentas reales de sidigital.com.mx. Estos 2 secrets YA EXISTÍAN
# desde el Proyecto 1 (el "cartero"), no hay que crear nada nuevo.
GMAIL_USER = os.environ["GMAIL_USER"]
GMAIL_PASSWORD = os.environ["GMAIL_PASSWORD"]
GMAIL_IMAP_HOST = "imap.gmail.com"
GMAIL_SMTP_HOST = "smtp.gmail.com"

# Las direcciones reales, usadas solo para RECONOCER de dónde venía
# cada correo reenviado (no para conectarse a ellas).
CUENTA_PRINCIPAL = "roberto.flores@sidigital.com.mx"
CUENTA_ML = "mercado-libre@sidigital.com.mx"

GEMINI_API_KEY = os.environ["GEMINI_API_KEY"]
GEMINI_MODEL = os.environ.get("GEMINI_MODEL", "gemini-3.1-flash-lite")

# A quién le llega el digest final (tu bandeja normal en Outlook).
DESTINATARIO_DIGEST = os.environ.get("DESTINATARIO_DIGEST", "roberto.flores@sidigital.com.mx")

# Cuántas horas hacia atrás se revisan (por default, 1 día completo).
HORAS_ATRAS = int(os.environ.get("HORAS_ATRAS", "24"))

# Dónde vive el Excel de licitaciones dentro del repo.
EXCEL_PATH = os.environ.get("EXCEL_PATH", "Licitaciones.xlsx")

# Palabras clave que hacen que una licitación merezca análisis a fondo.
PALABRAS_CLAVE_GIRO = [
    "fotocopiado", "fotocopiadora", "impresión", "impresion", "multifuncional",
    "escaneo", "escáner", "escaner", "arrendamiento de equipo", "tóner", "toner",
    "copiadora", "consumibles de impresión", "consumibles de impresion",
]

# Estados donde Roberto SÍ da servicio. Se puede editar esta lista libremente.
ESTADOS_PERMITIDOS = [
    "Ciudad de México", "CDMX", "Estado de México", "Edomex", "Puebla",
    "Querétaro", "Queretaro", "Hidalgo", "Morelos", "Tlaxcala",
]

ESTADOS_MEXICO = [
    "Aguascalientes", "Baja California", "Baja California Sur", "Campeche",
    "Chiapas", "Chihuahua", "Ciudad de México", "Coahuila", "Colima",
    "Durango", "Estado de México", "Guanajuato", "Guerrero", "Hidalgo",
    "Jalisco", "Michoacán", "Morelos", "Nayarit", "Nuevo León", "Oaxaca",
    "Puebla", "Querétaro", "Quintana Roo", "San Luis Potosí", "Sinaloa",
    "Sonora", "Tabasco", "Tamaulipas", "Tlaxcala", "Veracruz", "Yucatán",
    "Zacatecas",
]

# Columnas del formato robusto de licitaciones (deben existir en el Excel;
# si el archivo no existe, el script lo crea con estas columnas).
COLUMNAS_LICITACIONES = [
    # verdes — del correo, siempre
    "Unidad Compradora", "Descripción", "Número", "Estado", "Difusión",
    "Límite Propuestas", "Tipo", "Fuente", "Liga",
    # amarillas — del PDF, si se puede (quedan vacías si no)
    "Junta de Aclaraciones", "Fecha de Fallo", "Carácter", "Modalidad",
    "Marca Detectada", "Menciona NOM", "De Nuestro Giro",
    # azules — análisis de 8 puntos, lo llena Gemini
    "Resumen Ejecutivo", "Qué Piden", "Domicilios de Instalación",
    "Equipo y Cantidad", "Modalidad de Pago", "Specs Técnicas",
    "Soporte y SLA", "Vigencia del Contrato",
    # grises — las llena Roberto a mano
    "Decisión", "Motivo", "Asignado a",
]


# ===========================================================
# UTILIDADES GENERALES
# ===========================================================

def log(mensaje):
    """Imprime con fecha y hora, para que se vea claro en los logs de
    GitHub Actions."""
    print(f"[{datetime.now():%Y-%m-%d %H:%M:%S}] {mensaje}")


def decodificar_encabezado(valor):
    """Los asuntos y remitentes a veces vienen codificados raro
    (ej. '=?UTF-8?B?...?='). Esta función los deja en texto normal."""
    if not valor:
        return ""
    partes = decode_header(valor)
    resultado = ""
    for texto, codificacion in partes:
        if isinstance(texto, bytes):
            resultado += texto.decode(codificacion or "utf-8", errors="ignore")
        else:
            resultado += texto
    return resultado


def conectar_gmail():
    """Abre una conexión IMAP al Gmail 'cartero' y selecciona la
    bandeja de entrada. Aquí es donde llegan las copias reenviadas
    de las 2 cuentas reales de sidigital.com.mx."""
    conexion = imaplib.IMAP4_SSL(GMAIL_IMAP_HOST, 993)
    conexion.login(GMAIL_USER, GMAIL_PASSWORD)
    conexion.select("INBOX")
    return conexion


def buscar_correos_recientes(conexion, horas_atras=HORAS_ATRAS):
    """Devuelve una lista de mensajes (objetos email.message.Message)
    recibidos en las últimas `horas_atras` horas."""
    fecha_desde = (datetime.now() - timedelta(hours=horas_atras)).strftime("%d-%b-%Y")
    criterio = f'(SINCE "{fecha_desde}")'
    typ, datos = conexion.search(None, criterio)
    ids = datos[0].split()

    mensajes = []
    for id_correo in ids:
        typ, datos_msg = conexion.fetch(id_correo, "(RFC822)")
        if typ != "OK" or not datos_msg or not datos_msg[0]:
            continue
        mensajes.append(email.message_from_bytes(datos_msg[0][1]))
    return mensajes


def obtener_cuerpo_html_y_texto(mensaje):
    """De un correo (posiblemente multipart), extrae el HTML (si hay)
    y el texto plano (si hay)."""
    html = ""
    texto = ""
    if mensaje.is_multipart():
        for parte in mensaje.walk():
            tipo = parte.get_content_type()
            disposicion = str(parte.get("Content-Disposition") or "")
            if "attachment" in disposicion:
                continue
            try:
                contenido = parte.get_payload(decode=True)
                if contenido is None:
                    continue
                contenido = contenido.decode(parte.get_content_charset() or "utf-8", errors="ignore")
            except Exception:
                continue
            if tipo == "text/html":
                html += contenido
            elif tipo == "text/plain":
                texto += contenido
    else:
        try:
            contenido = mensaje.get_payload(decode=True)
            if contenido:
                contenido = contenido.decode(mensaje.get_content_charset() or "utf-8", errors="ignore")
                if mensaje.get_content_type() == "text/html":
                    html = contenido
                else:
                    texto = contenido
        except Exception:
            pass
    return html, texto


# ===========================================================
# PARTE 0 — DETECTAR DE CUÁL CUENTA VENÍA EL CORREO (NUEVO)
# ===========================================================
# Cuando cPanel reenvía un correo, conserva los encabezados originales
# de "a quién iba dirigido" (aunque el sobre SMTP apunte a Gmail).
# Buscamos esa dirección original en varios encabezados posibles.

def detectar_cuenta_origen(mensaje):
    """Devuelve 'principal', 'ml', o 'desconocido' según de cuál
    cuenta de sidigital.com.mx venía originalmente el correo."""
    encabezados_a_revisar = ["Delivered-To", "X-Original-To", "To", "Envelope-To"]
    texto_encabezados = ""
    for nombre in encabezados_a_revisar:
        valor = mensaje.get(nombre, "")
        texto_encabezados += decodificar_encabezado(valor).lower() + " "

    # Como respaldo, también revisamos la cadena "Received" (a veces
    # ahí queda registrado el destinatario original del reenvío).
    for valor in mensaje.get_all("Received", []):
        texto_encabezados += str(valor).lower() + " "

    if CUENTA_PRINCIPAL in texto_encabezados:
        return "principal"
    if CUENTA_ML in texto_encabezados:
        return "ml"
    return "desconocido"


# ===========================================================
# PARTE 1 — CLASIFICACIÓN DE LOS CORREOS DE LA CUENTA PRINCIPAL
# ===========================================================

REMITENTES_LICITACIONES = ["licitary.mx", "comprasmx", "compras mx", "alertas@licitary"]


def clasificar_correo_principal(mensaje):
    """Devuelve una de: 'licitaciones', 'clientes_proveedores', 'otros'."""
    remitente = decodificar_encabezado(mensaje.get("From", "")).lower()
    asunto = decodificar_encabezado(mensaje.get("Subject", "")).lower()

    if any(clave in remitente for clave in REMITENTES_LICITACIONES) or "licitac" in asunto:
        return "licitaciones"

    # Heurística simple para separar correos automáticos/genéricos de
    # correos de una persona real (cliente o proveedor).
    patrones_automaticos = [
        "noreply", "no-reply", "notificacion", "notification", "factura",
        "cfdi", "newsletter", "boletin", "aviso automático",
    ]
    if any(patron in remitente for patron in patrones_automaticos):
        return "otros"

    return "clientes_proveedores"


# ===========================================================
# PARTE 2 — PARSEO DE LA TABLA DE LICITACIONES (del correo, gratis)
# ===========================================================

def deducir_estado(unidad_compradora):
    """Busca el nombre de un estado mexicano dentro del texto de la
    unidad compradora. Si no encuentra ninguno, regresa 'Por confirmar'
    (típico de dependencias federales: IMSS, CFE, SEDENA, etc.)."""
    texto = unidad_compradora or ""
    for estado in ESTADOS_MEXICO:
        if estado.lower() in texto.lower():
            return estado
    return "Por confirmar"


def cumple_filtro_de_interes(descripcion, estado):
    """Decide si esta licitación merece el análisis a fondo con Gemini."""
    descripcion_lower = (descripcion or "").lower()
    tiene_palabra_clave = any(clave in descripcion_lower for clave in PALABRAS_CLAVE_GIRO)
    estado_ok = estado in ESTADOS_PERMITIDOS or estado == "Por confirmar"
    return tiene_palabra_clave and estado_ok


def parsear_tabla_licitaciones(html, fuente_default="Licitary/Compras MX"):
    """Busca en el HTML del correo una tabla con columnas tipo:
    Unidad Compradora | Descripción | Número | Difusión | Propuestas |
    Documentos | Fuente
    y regresa una lista de diccionarios, uno por licitación encontrada."""
    if not html:
        return []

    soup = BeautifulSoup(html, "html.parser")
    filas_resultado = []

    for tabla in soup.find_all("table"):
        encabezados = [th.get_text(strip=True).lower() for th in tabla.find_all(["th"])]
        filas = tabla.find_all("tr")
        if not filas:
            continue

        # Si no hay <th>, probamos usar la primera fila como encabezado.
        if not encabezados:
            primera = filas[0]
            encabezados = [c.get_text(strip=True).lower() for c in primera.find_all(["td", "th"])]
            filas = filas[1:]

        if not encabezados or "unidad" not in " ".join(encabezados):
            continue  # no es la tabla que buscamos

        for fila in filas:
            celdas = fila.find_all("td")
            if not celdas or len(celdas) < 2:
                continue

            dato = {}
            liga_encontrada = ""
            for i, celda in enumerate(celdas):
                nombre_columna = encabezados[i] if i < len(encabezados) else f"col_{i}"
                dato[nombre_columna] = celda.get_text(strip=True)
                enlace = celda.find("a")
                if enlace and enlace.get("href") and not liga_encontrada:
                    liga_encontrada = enlace["href"]

            unidad = dato.get("unidad compradora", "")
            if not unidad:
                continue

            filas_resultado.append({
                "Unidad Compradora": unidad,
                "Descripción": dato.get("descripción", dato.get("descripcion", "")),
                "Número": dato.get("número", dato.get("numero", "")),
                "Difusión": dato.get("difusión", dato.get("difusion", "")),
                "Límite Propuestas": dato.get("propuestas", ""),
                "Tipo": "Electrónica" if "electr" in html.lower() else "",
                "Fuente": dato.get("fuente", fuente_default),
                "Liga": liga_encontrada,
            })

    return filas_resultado


# ===========================================================
# PARTE 3 — ANÁLISIS PROFUNDO CON GEMINI (solo para las que interesan)
# ===========================================================

PROMPT_8_PUNTOS = """Actúa como experto analista en licitaciones públicas mexicanas.
Te paso el texto de las bases de una licitación (convocatoria y/o anexos
técnicos). Extrae la información en las 8 categorías de abajo, de forma
clara y directa. Si algo no aparece en el texto, escribe "No especificado".

Responde ÚNICAMENTE con un JSON válido, sin texto antes ni después, con
estas claves exactas:

{
  "resumen_ejecutivo": "fechas críticas: límite de entrega y apertura de
    propuestas (física o electrónica), fecha del fallo, vigencia del contrato",
  "que_piden": "descripción del servicio o bienes, quién pone los
    consumibles, modalidad de pago (consumo exacto o cuota fija)",
  "domicilios_instalacion": "institución contratante y direcciones exactas
    (calle, número, colonia/municipio) de cada unidad donde se instala/entrega",
  "equipo_y_cantidad": "tabla en texto: Ubicación | Cantidad y equipo |
    Características y especificaciones mínimas (usa viñetas con -)",
  "modalidad_de_pago": "resumen breve de la modalidad de pago",
  "specs_tecnicas": "cantidad total de equipos y volumen operativo estimado
    (ej. copias o servicios proyectados)",
  "soporte_y_sla": "técnicos de planta, tiempo límite de atención a fallas,
    política de sustitución de equipos, condiciones de mesa de ayuda",
  "vigencia_del_contrato": "requisitos legales: NOMs, licencias
    (municipales/COFEPRIS), o partnerships tecnológicos exigidos"
}

TEXTO DE LA LICITACIÓN:
---
{texto}
---
"""


def extraer_texto_pdf(contenido_bytes):
    """Recibe los bytes de un PDF y regresa el texto de todas sus
    páginas, unido en un solo string."""
    import io
    lector = PdfReader(io.BytesIO(contenido_bytes))
    texto_completo = []
    for pagina in lector.pages:
        try:
            texto_completo.append(pagina.extract_text() or "")
        except Exception:
            continue
    return "\n".join(texto_completo)


def descargar_pdf_si_aplica(liga):
    """Si la liga apunta directo a un PDF, lo descarga y regresa los
    bytes. Si no (ej. es una página dinámica de Compras MX con '#/'),
    regresa None."""
    if not liga:
        return None
    if "#/" in liga:
        return None  # portal dinámico, no se puede bajar directo (frágil)
    try:
        respuesta = requests.get(liga, timeout=30)
        content_type = respuesta.headers.get("Content-Type", "")
        if respuesta.status_code == 200 and (
            liga.lower().endswith(".pdf") or "pdf" in content_type.lower()
        ):
            return respuesta.content
    except Exception as error:
        log(f"  No se pudo descargar la liga {liga}: {error}")
    return None


def analizar_licitacion_con_gemini(texto_pdf, cliente_gemini):
    """Manda el texto del PDF a Gemini y regresa un diccionario con las
    8 categorías. Si algo falla, regresa un diccionario vacío."""
    try:
        # Los PDFs de licitaciones pueden ser largos; recortamos a un
        # tamaño razonable para no gastar de más.
        texto_recortado = texto_pdf[:400_000]
        prompt = PROMPT_8_PUNTOS.replace("{texto}", texto_recortado)

        respuesta = cliente_gemini.models.generate_content(
            model=GEMINI_MODEL,
            contents=prompt,
            config={"response_mime_type": "application/json"},
        )
        return json.loads(respuesta.text)
    except Exception as error:
        log(f"  Gemini no pudo analizar esta licitación: {error}")
        return {}


# ===========================================================
# PARTE 4 — EXCEL DE LICITACIONES
# ===========================================================

def abrir_o_crear_excel(ruta):
    if os.path.exists(ruta):
        libro = load_workbook(ruta)
        hoja = libro.active
        encabezados_actuales = [c.value for c in hoja[1]]
        # Si al Excel de Roberto le faltan columnas nuevas, las agregamos
        # al final sin tocar lo que ya tiene.
        for columna in COLUMNAS_LICITACIONES:
            if columna not in encabezados_actuales:
                hoja.cell(row=1, column=hoja.max_column + 1, value=columna)
        return libro, hoja
    else:
        libro = Workbook()
        hoja = libro.active
        hoja.title = "Licitaciones"
        hoja.append(COLUMNAS_LICITACIONES)
        return libro, hoja


def agregar_fila_licitacion(hoja, datos_fila):
    """Agrega una fila nueva, respetando el orden real de columnas que
    tenga la hoja (por si el Excel de Roberto ya trae otro orden)."""
    encabezados = [c.value for c in hoja[1]]
    fila_nueva = [datos_fila.get(col, "") for col in encabezados]
    hoja.append(fila_nueva)


# ===========================================================
# PARTE 5 — REPORTE DE MERCADO LIBRE (temporal, por correo)
# ===========================================================
# NOTA: esto es un reporte básico mientras se conecta la API oficial de
# Mercado Libre en el Paso 4. Por ahora solo cuenta por palabras clave
# en el asunto, sin poder ver el texto completo de cada pregunta.

PATRONES_ML = {
    "preguntas": ["pregunta", "te han preguntado", "nueva pregunta"],
    "ventas": ["venta", "vendiste", "nueva venta", "compraron"],
    "reclamos": ["reclamo", "mediación", "mediacion", "disputa"],
    "mensajes": ["mensaje", "te escribió", "te escribio"],
}


def clasificar_correo_ml(asunto):
    asunto_lower = (asunto or "").lower()
    for categoria, patrones in PATRONES_ML.items():
        if any(patron in asunto_lower for patron in patrones):
            return categoria
    return "otros"


def generar_reporte_ml(mensajes):
    conteos = {"preguntas": 0, "ventas": 0, "reclamos": 0, "mensajes": 0, "otros": 0}
    detalle = []
    for mensaje in mensajes:
        asunto = decodificar_encabezado(mensaje.get("Subject", ""))
        categoria = clasificar_correo_ml(asunto)
        conteos[categoria] += 1
        if categoria in ("preguntas", "reclamos"):
            detalle.append((categoria, asunto))
    return conteos, detalle


# ===========================================================
# PARTE 6 — ARMADO Y ENVÍO DEL DIGEST
# ===========================================================

def armar_html_digest(clientes_proveedores, otros, licitaciones_nuevas, reporte_ml):
    hoy = datetime.now().strftime("%d/%m/%Y")
    conteos_ml, detalle_ml = reporte_ml

    filas_licitaciones = ""
    for lic in licitaciones_nuevas:
        filas_licitaciones += f"""
        <tr>
          <td>{lic.get('Unidad Compradora', '')}</td>
          <td>{lic.get('Descripción', '')}</td>
          <td>{lic.get('Número', '')}</td>
          <td>{lic.get('Estado', '')}</td>
          <td>{lic.get('Límite Propuestas', '')}</td>
        </tr>"""

    detalle_ml_html = "".join(f"<li>[{c}] {a}</li>" for c, a in detalle_ml) or "<li>Sin novedades urgentes.</li>"

    lista_clientes = "".join(
        f"<li><b>{decodificar_encabezado(m.get('From',''))}</b> — {decodificar_encabezado(m.get('Subject',''))}</li>"
        for m in clientes_proveedores
    ) or "<li>Sin correos nuevos de clientes/proveedores.</li>"

    lista_otros = "".join(
        f"<li>{decodificar_encabezado(m.get('Subject',''))}</li>" for m in otros
    ) or "<li>Nada nuevo.</li>"

    return f"""
    <html><body style="font-family:sans-serif;">
      <h2>📬 Resumen de correos — {hoy}</h2>

      <h3>🛒 Mercado Libre</h3>
      <p>
        Preguntas: <b>{conteos_ml['preguntas']}</b> ·
        Ventas: <b>{conteos_ml['ventas']}</b> ·
        Reclamos: <b>{conteos_ml['reclamos']}</b> ·
        Mensajes: <b>{conteos_ml['mensajes']}</b>
      </p>
      <ul>{detalle_ml_html}</ul>
      <p style="color:#888;font-size:12px;">
        (Reporte básico por correo. En el Paso 4 se conecta la API oficial
        de Mercado Libre para un reporte más completo.)
      </p>

      <h3>🏛️ Licitaciones nuevas ({len(licitaciones_nuevas)})</h3>
      <table border="1" cellpadding="6" cellspacing="0">
        <tr><th>Unidad</th><th>Descripción</th><th>Número</th><th>Estado</th><th>Límite</th></tr>
        {filas_licitaciones or '<tr><td colspan="5">Sin licitaciones nuevas hoy.</td></tr>'}
      </table>
      <p style="color:#888;font-size:12px;">
        Ya se agregaron al Excel. Las marcadas de tu interés ya incluyen
        el análisis de Gemini en las columnas correspondientes.
      </p>

      <h3>🏢 Clientes / Proveedores</h3>
      <ul>{lista_clientes}</ul>

      <h3>📧 Otros / Informativos</h3>
      <ul>{lista_otros}</ul>

      <p style="color:#888;font-size:11px;margin-top:20px;">
        Este correo se generó automáticamente leyendo, vía reenvío, tus
        cuentas roberto.flores@ y mercado-libre@sidigital.com.mx.
      </p>
    </body></html>
    """


def enviar_digest(html):
    mensaje = MIMEMultipart("alternative")
    mensaje["Subject"] = f"📬 Tu resumen de correos — {datetime.now():%d/%m/%Y}"
    mensaje["From"] = GMAIL_USER
    mensaje["To"] = DESTINATARIO_DIGEST
    mensaje.attach(MIMEText(html, "html"))

    contexto = ssl.create_default_context()
    with smtplib.SMTP_SSL(GMAIL_SMTP_HOST, 465, context=contexto) as servidor:
        servidor.login(GMAIL_USER, GMAIL_PASSWORD)
        servidor.sendmail(GMAIL_USER, [DESTINATARIO_DIGEST], mensaje.as_string())


# ===========================================================
# PROGRAMA PRINCIPAL
# ===========================================================

def main():
    log("Iniciando clasificador de correos (vía Gmail)...")
    cliente_gemini = genai.Client(api_key=GEMINI_API_KEY)

    # --- Una sola conexión: al Gmail cartero ---
    log("Conectando al Gmail cartero...")
    imap = conectar_gmail()
    correos = buscar_correos_recientes(imap)
    imap.logout()
    log(f"  {len(correos)} correos encontrados en total.")

    # --- Separar por cuenta de origen ---
    correos_principal, correos_ml, correos_sin_identificar = [], [], []
    for correo in correos:
        origen = detectar_cuenta_origen(correo)
        if origen == "principal":
            correos_principal.append(correo)
        elif origen == "ml":
            correos_ml.append(correo)
        else:
            correos_sin_identificar.append(correo)

    if correos_sin_identificar:
        log(f"  Aviso: {len(correos_sin_identificar)} correos no se pudieron identificar "
            f"de qué cuenta venían (se ignoran). Revisar encabezados si esto crece mucho.")

    log(f"  {len(correos_principal)} de la cuenta principal, {len(correos_ml)} de Mercado Libre.")

    # --- Clasificación de la cuenta principal ---
    clientes_proveedores, otros, correos_licitaciones = [], [], []
    for correo in correos_principal:
        categoria = clasificar_correo_principal(correo)
        if categoria == "licitaciones":
            correos_licitaciones.append(correo)
        elif categoria == "clientes_proveedores":
            clientes_proveedores.append(correo)
        else:
            otros.append(correo)

    # --- Parseo de licitaciones + análisis con Gemini ---
    log(f"Procesando {len(correos_licitaciones)} correos de licitaciones...")
    libro_excel, hoja_excel = abrir_o_crear_excel(EXCEL_PATH)
    licitaciones_nuevas = []

    for correo in correos_licitaciones:
        html, _ = obtener_cuerpo_html_y_texto(correo)
        filas = parsear_tabla_licitaciones(html)
        for fila in filas:
            fila["Estado"] = deducir_estado(fila["Unidad Compradora"])
            merece_analisis = cumple_filtro_de_interes(fila["Descripción"], fila["Estado"])
            fila["De Nuestro Giro"] = "Sí" if merece_analisis else ""

            if merece_analisis:
                log(f"  Licitación de interés: {fila['Descripción'][:60]}...")
                pdf_bytes = descargar_pdf_si_aplica(fila.get("Liga", ""))
                if pdf_bytes:
                    texto_pdf = extraer_texto_pdf(pdf_bytes)
                    analisis = analizar_licitacion_con_gemini(texto_pdf, cliente_gemini)
                    fila["Resumen Ejecutivo"] = analisis.get("resumen_ejecutivo", "")
                    fila["Qué Piden"] = analisis.get("que_piden", "")
                    fila["Domicilios de Instalación"] = analisis.get("domicilios_instalacion", "")
                    fila["Equipo y Cantidad"] = analisis.get("equipo_y_cantidad", "")
                    fila["Modalidad de Pago"] = analisis.get("modalidad_de_pago", "")
                    fila["Specs Técnicas"] = analisis.get("specs_tecnicas", "")
                    fila["Soporte y SLA"] = analisis.get("soporte_y_sla", "")
                    fila["Vigencia del Contrato"] = analisis.get("vigencia_del_contrato", "")
                else:
                    fila["Resumen Ejecutivo"] = "Requiere revisión manual (liga dinámica o sin PDF directo)."

            agregar_fila_licitacion(hoja_excel, fila)
            licitaciones_nuevas.append(fila)

    libro_excel.save(EXCEL_PATH)
    log(f"Excel actualizado: {EXCEL_PATH} ({len(licitaciones_nuevas)} filas nuevas).")

    # --- Reporte de Mercado Libre ---
    reporte_ml = generar_reporte_ml(correos_ml)
    log(f"  Reporte ML: {reporte_ml[0]}")

    # --- Armado y envío del digest ---
    log("Armando y enviando el digest...")
    html_digest = armar_html_digest(clientes_proveedores, otros, licitaciones_nuevas, reporte_ml)
    enviar_digest(html_digest)
    log("¡Listo! Digest enviado.")


if __name__ == "__main__":
    main()
